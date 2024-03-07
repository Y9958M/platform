
import logging
import colorlog
import os
import time
import threading
import itertools
import decimal
import json
import redis
import sqlalchemy
import requests
from logging.handlers import RotatingFileHandler
from datetime import date, datetime
from sqlalchemy.pool import QueuePool
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import importlib
import platform
import pandas as pd
# import numpy as np

s_os_path = 'D:/home' if platform.system() == 'Windows' else '/home'

CF = importlib.machinery.SourceFileLoader('config', f'{s_os_path}/platform/config.py').load_module()

# & d:/CODE/platform/.venv/Scripts/pip.exe install openpyxl
# 要提供给其它项目调用的共用方法
DB_LINK = CF.DB_LINK if isinstance(CF.DB_LINK,dict) else {}
PF_LINK = CF.PF_LINK if isinstance(CF.PF_LINK,dict) else {}
SID = CF.SID if isinstance(CF.SID,int) else 0
ADMIN = CF.ADMIN if isinstance(CF.ADMIN,list) else []

CLIENT ="Dev:10.56"
VER = 240201
PROJECT = "YM"
MESSAGE = {
    "code": 500,
    "sid":  SID,
    "count":  0,
    "msg":  '',
    "project":PROJECT,
    "client":CLIENT,
    "ver":VER,
    "author":'姚鸣'
}
LTD = "Copyright© 2023 by SH-Mart"
IGNORE = {'id','ldt','cdt'}

logging._srcfile = None
logging.logThreads = 0
logging.logMultiprocessing = 0
logging.logProcesses = 0
logging.thread = None  # type: ignore
log_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), 'logs')  # log_path为存放日志的路径
if not os.path.exists(log_path): os.mkdir(log_path)  # 若不存在logs文件夹，则自动创建  # noqa: E701
s_log_file = os.path.join(log_path,f"mp-{datetime.now().strftime('%Y%m%d')}.log")


rs = redis.Redis(connection_pool=redis.ConnectionPool(
    host    =   DB_LINK['REDIS']['HOST'],
    password =  DB_LINK['REDIS']['PWD'],
    port =      DB_LINK['REDIS']['PORT'],
    db =        DB_LINK['REDIS']['DB'],
    decode_responses=True,
    encoding='utf-8'))


# 初始化配置 数据从后台取 初始化的时候 取项目库JSON数据# 连接后台 启动级异常 不需要返回 直接记日志
class HandleLog:
    def __init__(self,s_name, s_path=s_log_file, i_c_level = 60 - SID*10, i_f_level = 60 - SID*10):
        self.logger = logging.getLogger(s_name) # 创建日志记录器
        self.logger.setLevel(logging.DEBUG)
        # self.__logger = logging.getLogger() 
        formatter = colorlog.ColoredFormatter(
        '%(log_color)s%(asctime)s %(name)6s: %(message_log_color)s%(message)s',
        datefmt='%y%m%d %H:%M:%S',
        reset=True,
        log_colors={
            'DEBUG':    'black',
            'INFO':     'cyan',
            'WARNING':  'yellow',
            'ERROR':    'red',
            'CRITICAL': 'red,bg_white',
        },
        secondary_log_colors={'message':{
            'DEBUG':    'light_black',
            'INFO':     'light_cyan',
            'WARNING':  'light_yellow',
            'ERROR':    'light_red',
            'CRITICAL': 'light_purple'
        }},
        style='%')

        #设置CMD日志
        sh = logging.StreamHandler()
        sh.setFormatter(formatter)
        sh.setLevel(i_c_level)

        #设置文件日志
        fh = RotatingFileHandler(filename=s_path, mode="w", maxBytes=5*1024*1024, backupCount=5,encoding='utf-8')
        formatter_file = logging.Formatter('%(asctime)s %(name)s %(levelname)9s: %(message)s', datefmt='%a %y%m%d %H:%M:%S')
        fh.setFormatter(formatter_file)
        fh.setLevel(i_f_level)

        self.logger.propagate = False
        if not self.logger.handlers: # 防止日志重复打印 logger.propagate 布尔标志, 用于指示消息是否传播给父记录器
            self.logger.addHandler(sh)
            self.logger.addHandler(fh)
    
    def debug(self,message,title=''):
        self.logger.debug(f"{title}:{message}" if title else message)

    def info(self,message,title=''):
        self.logger.info(f"{title}:{message}" if title else message)

    def warning(self,message,title=''):
        self.logger.warning(f"{title}:{message}" if title else message)

    def error(self,message,title=''):
        self.logger.error(f"{title}:{message}" if title else message)

    def cri(self,message,title=''):
        self.logger.critical(f"{title}:{message}" if title else message)

log = HandleLog("main")

def setGlobal()->dict:
    s_conn = f"mysql+pymysql://{DB_LINK['YM']['USER']}:{DB_LINK['YM']['PWD']}@{DB_LINK['YM']['HOST']}:{DB_LINK['YM']['PORT']}/platform"
    log.debug(s_conn,'s_conn')
    s_ = "SELECT json_values FROM set_global WHERE project_name = 'GLOBAL' AND project_key = 'SETUP';"
    try:
        df = pd.read_sql_query(s_,sqlalchemy.create_engine(s_conn))['json_values'].head(1)
        if df.shape[0] == 0:
            raise Exception("SELECT json_values FROM set_global WHERE project_name = 'GLOBAL' AND project_key = 'SETUP'")
        else:
            j_ = eval(df.to_dict()[0])
    except Exception as e:
        log.error(e)
    if 'VER' not in j_.keys():
        raise Exception("Yao Ming tell you: set_global > SETUP > json_values > VER is NEED! ")
    elif j_['VER'] != VER:
        raise Exception(f"Yao Ming tell you: PROJECT:{VER} DB:{j_['VER']} VER IS NOT MARRY ! ")
    else:
        print(f"{'-' * 24} platform {'-' * 24}")
        print(j_)
    return j_

# 反转字典
def reverse_dict(input_dict):
    return dict(zip(input_dict.values(), input_dict.keys()))


def l2d(j_ds):
    if 'data' not in j_ds:
        return j_ds
    if 'fields' not in j_ds['data'] or 'datalist' not in j_ds['data']:
        return j_ds
    fields = j_ds['data']['fields']
    datalist = j_ds['data']['datalist']
    l_ds = [dict(zip(fields, sublist)) for sublist in datalist]
    j_ds['data']['datalist'] = l_ds
    return j_ds


def engine(DB='platform',LINK=""):
    LINK = LINK if LINK else 'YM'
    JDBC ={'MYSQL':'mysql+pymysql','MSSQL':'mssql+pymssql'}
    PARM ={'MYSQL':'','MSSQL':'?charset=cp936'}
    s_ = f"{JDBC[DB_LINK[LINK]['TYPE']]}://{DB_LINK[LINK]['USER']}:{DB_LINK[LINK]['PWD']}@{DB_LINK[LINK]['HOST']}:{DB_LINK[LINK]['PORT']}/{DB}{PARM[DB_LINK[LINK]['TYPE']]}"
    return sqlalchemy.create_engine(s_, poolclass=QueuePool, max_overflow=10, pool_size=24, pool_timeout=30, pool_recycle=3600)


# -*- 把Date、DateTime类型数据转换成兼容Json的格式 -*-  json.dumps(result,cls=DateEncoder.DateEncoder) # 调用自定义类
class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(obj, date):
            return obj.strftime("%Y-%m-%d")
        elif isinstance(obj, int):
            return int(obj)
        elif isinstance(obj, decimal.Decimal):
            return float(obj)
        elif isinstance(obj,sqlalchemy.engine.row.Row):
            return list(obj)
        elif isinstance(obj,sqlalchemy.engine.row.RowMapping):
            return dict(obj)
        elif isinstance(obj,sqlalchemy.engine.result.RMKeyView):
            return list(obj)
        else:
            return json.JSONEncoder.default(self, obj)

def msgJson(data):
    return json.dumps(data,cls=DateEncoder,ensure_ascii=False)

def msgWrapper(ldt:int,s_func_remark=''):
    j_msg = MESSAGE.copy()
    def reMsg(func):
        def wrapped_function(*args, **kwargs):
            j_msg['Fac'] = func.__name__
            start_time = time.time()
            start_strftime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

            j_res = func(*args, **kwargs)

            end_strftime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            d_time = time.time() - start_time
            j_msg.update(j_res)
            j_msg.update({'start_time':start_strftime,'end_time':end_strftime,'times':round(d_time,2),"ldt":ldt,'func_remark':s_func_remark})
            return j_msg
        return wrapped_function
    return reMsg

# 要区分一下 查询 JOB 单据 job_custom_logs 
class threadLogs(threading.Thread):
    def __init__(self,thread_id,thread_name:str,fac:str,args_dict={}):
        threading.Thread.__init__(self)
        self.thread_id = thread_id
        self.thread_name = thread_name
        self.args_dict = args_dict
        self.fac = fac

    def run(self):
        FAC = self.fac
        LOGS = {
    "commonQuery":    "insert into log_common_query(userid,sqlid,parm_json)values(%s,%s,%s)",
    "commonUpdate":   "insert into log_common_update(userid,sqlid,parm_json)values(%s,%s,%s)",
    "commonRedis":    "insert into log_common_redis(userid,redis_name,parm_json)values(%s,%s,%s)",
    "commonBill":     "insert into log_common_bill(userid,billid,parm_json)values(%s,%s,%s)",
    "authLogin":      "insert into log_auth_login(code_from,userid,parm_json)values(%s,%s,%s)",
    "postJob":        "insert into log_post_job(jobid,userid,parm_json)values(%s,%s,%s)",
    }
        time.sleep(3)
        conn = engine().connect()
        try:
            if FAC in LOGS:
                args = self.args_dict
                if FAC in ['commonQuery'] and 'data' in args:
                    if 'datalist' in args['data']:
                        args['tip'] ="3s del data_list!"
                        del args['data']['datalist']
                conn.exec_driver_sql(LOGS[FAC],(self.thread_id,self.thread_name,msgJson(args)))
                conn.commit()
            else:
                conn.exec_driver_sql("insert into log_def(threadid,thread_name,parm_json)values(%s,%s,%s)",
                (self.thread_id,self.thread_name,msgJson(self.args_dict)))
                conn.commit()
        except Exception as e:
            print("threadLogs",e)


# 发送文本消息
def send_text(webhook, content, mentioned_list=None, mentioned_mobile_list=None):
    header = {
                "Content-Type": "application/json",
                "Charset": "UTF-8"
                }
    data ={
        "msgtype": "text",
        "text": {
            "content": content
            ,"mentioned_list":mentioned_list
            ,"mentioned_mobile_list":mentioned_mobile_list
        }
    }
    data = json.dumps(data, ensure_ascii=False).encode("utf-8")
    info = requests.post(url=webhook, data=data, headers=header)
    print(info)
    

# 发送markdown消息
def send_md(webhook, content):
    header = {
                "Content-Type": "application/json",
                "Charset": "UTF-8"
                }
    data ={
        "msgtype": "markdown",
        "markdown": {
            "content": content
        }
    }
    # data = json.dumps(data)
    data = json.dumps(data, ensure_ascii=False).encode("utf-8")
    info = requests.post(url=webhook, data=data, headers=header)
    print(info)

# 发送文件
def send_file(webhook, file):
    # 获取media_id
    key = webhook.split('key=')[1]
    id_url = f'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file'
    files = {'file': open(file, 'rb')}
    res = requests.post(url=id_url, files=files)
    media_id = res.json()['media_id']
    
    header = {
                "Content-Type": "application/json",
                "Charset": "UTF-8"
                }
    data ={
    "msgtype": "file",
    "file": {
                "media_id": media_id
        }
    }
    # data = json.dumps(data)
    data = json.dumps(data, ensure_ascii=False).encode("utf-8")
    info = requests.post(url=webhook, data=data, headers=header)
    print(info)

# 生成EXCEL
def df2Excel(df:pd.core.frame.DataFrame,s_path:str,s_title:str,j_args={})->dict:
    log.debug(s_path,s_title)
    with pd.ExcelWriter(s_path,mode='w',date_format="YYYY-MM-DD",datetime_format="YYYY-MM-DD HH:MM:SS",engine='openpyxl') as writer:
        df.to_excel(writer,sheet_name=s_title,index=False,freeze_panes=(1,0),startrow=0,na_rep="",float_format ="%.6f")
        j_char = {1:'A',2:'B',3:'C',4:'D',5:'E',6:'F',7:'G',8:'H',9:'I',10:'J',11:'K',12:'L',13:'M',14:'N',15:'O',16:'P',17:'Q',18:'R',19:'S',20:'T',21:'U',22:'V',23:'W',24:'X',25:'Y',26:'Z'
        ,27:'AA',28:'AB',29:'AC',30:'AD',31:'AE',32:'AF',33:'AG',34:'AH',35:'AI',36:'AJ',37:'AK',38:'AL',39:'AM',40:'AN',41:'AO',42:'AP',43:'AQ',44:'AR',45:'AS',46:'AT',47:'AU',48:'AV',49:'AW',50:'AX',51:'AY',52:'AZ'
        ,53:'BA',54:'BB',55:'BC',56:'BD',57:'BE',58:'BF',59:'BG',60:'BH',61:'BI',62:'BJ',63:'BK',64:'BL',65:'BM',66:'BN',67:'BO',68:'BP',69:'BQ',70:'BR',71:'BS',72:'BT',73:'BU',74:'BV',75:'BW',76:'BX',77:'BY',78:'BZ'
        ,79:'CA',80:'CB',81:'CC',82:'CD',83:'CE',84:'CF',85:'CG',86:'CH',87:'CI',88:'CJ',89:'CK',90:'CL',91:'CM',92:'CN',93:'CO',94:'CP',95:'CQ',96:'CR',97:'CS',98:'CT',99:'CU',100:'CV',101:'CW',102:'CX',103:'CY',104:'CZ'}

        i_high = df.shape[0] + 1
        s_with = j_char[df.shape[1]]
        log.debug(i_high,'i_high')
        log.debug(s_with,'s_with')

        # workbook = writer.book
        font = Font(name="微软雅黑", bold=False)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        pattern_fill = PatternFill(fill_type="solid", fgColor="B4C6E7")
        side = Side(style="thin")
        border = Border(left=side, right=side, top=side, bottom=side)  

        ws = writer.sheets[s_title]
        try:
            for k,row in j_args.items():
                k = k.upper()
                if k not in j_char.values():
                    log.debug(k,'不在设置范围内:')
                    continue
                if 'width' in row:
                    ws.column_dimensions[k].width=row['width']
                if 'format' in row:
                    ws[f"{k}2:{k}{i_high}"].number_format = row['format']
        except Exception as e:
            log.error(e,'设置EXCEL格式出错')

        for cell in itertools.chain(*ws[f"A1:{s_with}1"]):
            cell.font = Font(name="微软雅黑", bold=True)
            cell.alignment = alignment
            cell.fill = pattern_fill
            cell.border = border

        for cell in itertools.chain(*ws[f"A2:{s_with}{i_high}"]):
            cell.font = font
            cell.alignment = alignment
            cell.border = border
    return {}

GLOBAL = setGlobal()
